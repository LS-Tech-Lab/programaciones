import os
import glob
import re
import pandas as pd
from google.colab import files
import openpyxl
from openpyxl.styles import PatternFill, Font
from collections import Counter
from tqdm.notebook import tqdm 
import time
import warnings

# Silenciamos advertencias estéticas de lectura de openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 1. Buscar de forma automática TODOS los archivos Excel (.xlsx) en la carpeta
archivos_excel = glob.glob("./*.xlsx")
archivos_validos = [f for f in archivos_excel if not os.path.basename(f).startswith("~$") and "Carga_" not in f]

if not archivos_validos:
    print("❌ No se encontraron archivos con extensión .xlsx para procesar.")
else:
    print(f"📚 Archivos detectados para procesamiento: {len(archivos_validos)}")
    datos_consolidados = []
    
    # BARRA DE PROGRESO 1: Avance general de archivos
    for archivo_excel in tqdm(archivos_validos, desc="Progreso Global de Archivos"):
        nombre_archivo = os.path.basename(archivo_excel)
        nombre_archivo_upper = nombre_archivo.upper()
        
        # DETECCIÓN ESTRATÉGICA DEL PROGRAMA BASE (Por nombre de archivo)
        programa_base = "OTROS"
        if any(x in nombre_archivo_upper for x in ["INGENIERIA", "INGENIERÍA", "ING."]):
            programa_base = "INGENIERÍA"
        elif any(x in nombre_archivo_upper for x in ["ADMINISTRACION", "ADMINISTRACIÓN", "ADMIN"]):
            programa_base = "ADMINISTRACIÓN"
        elif any(x in nombre_archivo_upper for x in ["EDUCACION", "EDUCACIÓN", "EDU."]):
            programa_base = "EDUCACIÓN"
        elif "PNF" in nombre_archivo_upper:
            programa_base = "PNF"
            
        excel_completo = pd.ExcelFile(archivo_excel)
        todas_las_hojas = excel_completo.sheet_names
        hojas_sedes = [h for h in todas_las_hojas if not any(x in h.upper() for x in ["PERMISOS", "INACTIVOS", "STATUS", "PLANTILLA", "CONSOLIDADO"])]
        
        # BARRA DE PROGRESO 2: Avance interno de pestañas por sede
        for hoja in tqdm(hojas_sedes, desc=f"↳ Procesando Sede (Libro: {nombre_archivo[:15]}...)", leave=False):
            
            # ANÁLISIS BASE DE LA SEDE (Por nombre de la hoja)
            nombre_hoja_upper = hoja.strip().upper()
            sede_inicial = hoja.strip() 
            
            if "LA CAÑADA" in nombre_hoja_upper: sede_inicial = "LA CAÑADA"
            elif "LOS PUERTOS" in nombre_hoja_upper: sede_inicial = "LOS PUERTOS"
            elif "OJEDA" in nombre_hoja_upper: sede_inicial = "CIUDAD OJEDA"
            elif "FRANCISCO" in nombre_hoja_upper or "SAN FRANCISCO" in nombre_hoja_upper: sede_inicial = "SAN FRANCISCO"
            elif "SANTA RITA" in nombre_hoja_upper: sede_inicial = "SANTA RITA"
            elif "MENE GRANDE" in nombre_hoja_upper: sede_inicial = "MENE GRANDE"
            elif "BOBURES" in nombre_hoja_upper: sede_inicial = "BOBURES"
            elif "CORO" in nombre_hoja_upper: sede_inicial = "CORO"
            elif "TRUJILLO" in nombre_hoja_upper: sede_inicial = "TRUJILLO"
            elif "CABIMAS" in nombre_hoja_upper: sede_inicial = "CABIMAS"
            elif "BACHAQUERO" in nombre_hoja_upper: sede_inicial = "BACHAQUERO"
            
            # CLASIFICACIÓN DE LA PESTAÑA
            programa_hoja = "OTROS"
            if any(x in nombre_hoja_upper for x in ["EDUCACION", "EDUCACIÓN", "HISTORIA", "ESPECIAL"]):
                programa_hoja = "EDUCACIÓN"
            elif any(x in nombre_hoja_upper for x in ["INGENIERIA", "INGENIERÍA"]):
                programa_hoja = "INGENIERÍA"
            elif any(x in nombre_hoja_upper for x in ["ADMINISTRACION", "ADMINISTRACIÓN"]):
                programa_hoja = "ADMINISTRACIÓN"
            elif "PNF" in nombre_hoja_upper:
                for sub in ["PNFI", "PNFCP", "PNFEEE", "PNFAGRO", "PNFH", "PNFEE"]:
                    if sub in nombre_hoja_upper: programa_hoja = sub
                if programa_hoja == "OTROS": programa_hoja = "PNF"
            
            df_completo = pd.read_excel(archivo_excel, sheet_name=hoja, dtype=str)
            
            fila_cabecera = None
            programa_celda = "OTROS"
            
            for idx, row in df_completo.astype(str).iterrows():
                fila_unida = " ".join(row.values).upper()
                
                if programa_base == "OTROS" and programa_hoja == "OTROS":
                    if any(x in fila_unida for x in ["EDUCACION", "EDUCACIÓN"]): programa_celda = "EDUCACIÓN"
                    elif any(x in fila_unida for x in ["INGENIERIA", "INGENIERÍA"]): programa_celda = "INGENIERÍA"
                    elif any(x in fila_unida for x in ["ADMINISTRACION", "ADMINISTRACIÓN"]): programa_celda = "ADMINISTRACIÓN"
                    elif "PNFI" in fila_unida: programa_celda = "PNFI"
                    elif "PNFCP" in fila_unida: programa_celda = "PNFCP"
                    elif "PNFEEE" in fila_unida: programa_celda = "PNFEEE"
                    elif "PNFAGRO" in fila_unida: programa_celda = "PNFAGRO"
                    elif "PNFH" in fila_unida: programa_celda = "PNFH"
                    elif "PNFEE" in fila_unida: programa_celda = "PNFEE"
                    elif "PNF" in fila_unida: programa_celda = "PNF"
                
                if "CÉDULA" in fila_unida or "APELLIDO" in fila_unida or "CEDULA" in fila_unida:
                    fila_cabecera = idx
                    break
            
            if programa_base != "OTROS":
                programa_inicial = programa_base
            elif programa_hoja != "OTROS":
                programa_inicial = programa_hoja
            else:
                programa_inicial = programa_celda
            
            if fila_cabecera is not None:
                df_sede = pd.read_excel(archivo_excel, sheet_name=hoja, skiprows=fila_cabecera + 1, dtype=str)
            else:
                df_sede = df_completo
                
            df_sede.columns = df_sede.columns.astype(str).str.strip().str.upper()
            
            c_ced = next((c for c in df_sede.columns if 'CÉDULA' in c or 'CEDULA' in c), None)
            c_nom = next((c for c in df_sede.columns if 'APELLIDO' in c or 'NOMBRE' in c), None)
            c_cor = next((c for c in df_sede.columns if 'CORREO' in c), None)
            c_tel = next((c for c in df_sede.columns if 'TELÉFONO' in c or 'TELEFONO' in c or 'TEL' in c), None)
            c_obs = next((c for c in df_sede.columns if 'OBSERVACIONES' in c or 'OBSERVACION' in c or 'OBS' in c), None)
            c_per = next((c for c in df_sede.columns if 'PERIODO' in c or 'PERÍODO' in c or 'LAPSO' in c), None)
            c_prog_col = next((c for c in df_sede.columns if 'PROGRAMA' in c or 'DEPENDENCIA' in c or 'CARRERA' in c), None)
            c_sede_col = next((c for c in df_sede.columns if 'SEDE' in c or 'UBICACIÓN' in c or 'UBICACION' in c), None)
            
            c_hrs = None
            for palabra in ['TOTAL DE HORAS', 'TOTAL CARGA ACADÉMICA', 'TOTAL', 'TOTAL HORAS', 'CARGA/HRS']:
                c_hrs = next((c for c in df_sede.columns if palabra in c), None)
                if c_hrs: break
            if not c_hrs:
                c_hrs = next((c for c in df_sede.columns if 'HORAS' in c), None)
                
            columnas_mapeadas = {}
            if c_ced: columnas_mapeadas[c_ced] = 'Cedula'
            if c_nom: columnas_mapeadas[c_nom] = 'Docente'
            if c_cor: columnas_mapeadas[c_cor] = 'Correo'
            if c_tel: columnas_mapeadas[c_tel] = 'Telefono'
            if c_obs: columnas_mapeadas[c_obs] = 'Observaciones'
            if c_per: columnas_mapeadas[c_per] = 'Periodo'
            if c_hrs: columnas_mapeadas[c_hrs] = 'Horas'
            if c_prog_col: columnas_mapeadas[c_prog_col] = 'Programa_Origen'
            if c_sede_col: columnas_mapeadas[c_sede_col] = 'Sede_Origen'
            
            if c_ced and c_nom:
                df_filtrado = df_sede[list(columnas_mapeadas.keys())].rename(columns=columnas_mapeadas)
                df_filtrado = df_filtrado.fillna('')
                
                # Extracción impecable de Cédulas reales
                def extraer_cedula_intacta(val):
                    val_str = str(val).strip()
                    if not val_str or val_str.lower() == 'nan': return ''
                    if val_str.endswith('.0'): 
                        val_str = val_str[:-2]
                    digitos = re.sub(r'\D', '', val_str)
                    if len(digitos) <= 4: return '' 
                    return digitos

                df_filtrado['Cedula'] = df_filtrado['Cedula'].apply(extraer_cedula_intacta)
                df_filtrado = df_filtrado[(df_filtrado['Cedula'] != '') | (df_filtrado['Docente'] != '')]
                
                # LÓGICA DE DETECCIÓN MULTITAREA PARA LA SEDE POR FILA
                def determinar_sede_fila(row_data, sede_init):
                    texto_celda_sede = str(row_data.get('Sede_Origen', '')).upper() + " " + str(row_data.get('Observaciones', '')).upper()
                    sedes_encontradas = []
                    
                    if "LA CAÑADA" in texto_celda_sede: sedes_encontradas.append("LA CAÑADA")
                    if "LOS PUERTOS" in texto_celda_sede: sedes_encontradas.append("LOS PUERTOS")
                    if "OJEDA" in texto_celda_sede: sedes_encontradas.append("CIUDAD OJEDA")
                    if "FRANCISCO" in texto_celda_sede or "SAN FRANCISCO" in texto_celda_sede: sedes_encontradas.append("SAN FRANCISCO")
                    if "SANTA RITA" in texto_celda_sede: sedes_encontradas.append("SANTA RITA")
                    if "MENE GRANDE" in texto_celda_sede: sedes_encontradas.append("MENE GRANDE")
                    if "BOBURES" in texto_celda_sede: sedes_encontradas.append("BOBURES")
                    if "CORO" in texto_celda_sede: sedes_encontradas.append("CORO")
                    if "TRUJILLO" in texto_celda_sede: sedes_encontradas.append("TRUJILLO")
                    if "CABIMAS" in texto_celda_sede: sedes_encontradas.append("CABIMAS")
                    if "BACHAQUERO" in texto_celda_sede: sedes_encontradas.append("BACHAQUERO")
                    
                    if sedes_encontradas:
                        # Eliminar duplicados manteniendo orden
                        sedes_unicas = list(dict.fromkeys(sedes_encontradas))
                        return " / ".join(sedes_unicas)
                    return sede_init

                df_filtrado['Sede'] = df_filtrado.apply(lambda r: determinar_sede_fila(r, sede_inicial), axis=1)
                
                # LÓGICA DE DETECCIÓN MULTITAREA DE PROGRAMAS EXTENDIDA (GLOBAL)
                def determinar_programa_fila(row_data, prog_inicial):
                    texto_celda = str(row_data.get('Programa_Origen', '')).upper() + " " + str(row_data.get('Observaciones', '')).upper()
                    
                    programas_detectados = []
                    
                    # 1. Escaneo Completo de Sub-Carreras PNF
                    if re.search(r'INFORMATICA|INFORMÁTICA|PNFI', texto_celda):
                        programas_detectados.append("PNFI")
                    if re.search(r'CONTADURIA|CONTADURÍA|PNFCP|CONTABLE|PÚBLICA|PUBLICA', texto_celda):
                        programas_detectados.append("PNFCP")
                    if re.search(r'ELECTRICIDAD|ELECTRONICA|ELECTRÓNICA|PNFEEE', texto_celda):
                        programas_detectados.append("PNFEEE")
                    if re.search(r'AGROALIMENTARIA|AGRO|PNFAGRO', texto_celda):
                        programas_detectados.append("PNFAGRO")
                    if re.search(r'HISTORIA|PNFH', texto_celda):
                        programas_detectados.append("PNFH")
                    if re.search(r'ESPECIAL|PNFEE', texto_celda):
                        programas_detectados.append("PNFEE")
                        
                    # 2. Escaneo Completo de Programas Tradicionales Independientes
                    if re.search(r'INGENIERIA|INGENIERÍA', texto_celda): 
                        programas_detectados.append("INGENIERÍA")
                    if re.search(r'ADMINISTRACION|ADMINISTRACIÓN', texto_celda): 
                        programas_detectados.append("ADMINISTRACIÓN")
                    if re.search(r'EDUCACION|EDUCACIÓN', texto_celda): 
                        programas_detectados.append("EDUCACIÓN")
                    
                    # Si la celda detalló programas de forma mixta, unificarlos dinámicamente
                    if programas_detectados:
                        programas_unicos = list(dict.fromkeys(programas_detectados))
                        return " / ".join(programas_unicos)
                    
                    # Si la celda vino vacía, heredar el contexto inicial estructural del archivo/pestaña
                    if prog_inicial in ["INGENIERÍA", "ADMINISTRACIÓN", "EDUCACIÓN", "PNFI", "PNFCP", "PNFEEE", "PNFAGRO", "PNFH", "PNFEE"]:
                        return prog_inicial
                        
                    if "PNF" in texto_celda:
                        return "PNF"
                        
                    return "PNF" if prog_inicial == "OTROS" else prog_inicial

                df_filtrado['Programa'] = df_filtrado.apply(lambda r: determinar_programa_fila(r, programa_inicial), axis=1)
                
                if 'Horas' in df_filtrado.columns:
                    df_filtrado['Horas'] = pd.to_numeric(df_filtrado['Horas'], errors='coerce').fillna(0)
                else:
                    df_filtrado['Horas'] = 0
                    
                if 'Periodo' not in df_filtrado.columns or df_filtrado['Periodo'].eq('').all():
                    df_filtrado['Periodo'] = 'I-2026'
                    
                # Función de Teléfonos Blindada
                def normalizar_telefono_estricto(val):
                    val_str = str(val).strip()
                    if not val_str or val_str.lower() == 'nan' or val_str == '0': return ''
                    if val_str.endswith('.0'): val_str = val_str[:-2]
                    if 'E+' in val_str.upper() or 'E' in val_str.upper():
                        try: val_str = str(int(float(val_str)))
                        except: pass
                    
                    sub_elementos = []
                    if '/' in val_str:
                        partes = val_str.split('/')
                        izq = re.sub(r'\D', '', partes[0])
                        der = re.sub(r'\D', '', partes[1])
                        if len(izq) <= 4 and len(der) >= 7: sub_elementos = [izq + der]
                        else: sub_elementos = partes
                    else:
                        sub_elementos = val_str.split(',') if ',' in val_str else [val_str]
                    
                    telefonos_procesados = []
                    for elemento in sub_elementos:
                        digitos = re.sub(r'\D', '', elemento)
                        if not digitos: continue
                        if len(digitos) == 11 and digitos.startswith('0'): telefonos_procesados.append(digitos)
                        elif len(digitos) == 10 and digitos.startswith(('4', '2')): telefonos_procesados.append('0' + digitos)
                        elif len(digitos) == 12 and digitos.startswith('58'): telefonos_procesados.append('0' + digitos[2:])
                        elif len(digitos) == 7: telefonos_procesados.append(digitos)
                        elif len(digitos) >= 7: telefonos_procesados.append(digitos)
                        
                    telefonos_unicos = []
                    for t in telefonos_procesados:
                        if t not in telefonos_unicos: telefonos_unicos.append(t)
                    return " / ".join(telefonos_unicos)

                for col in ['Docente', 'Correo', 'Telefono', 'Observaciones', 'Periodo']:
                    if col in df_filtrado.columns:
                        if col == 'Telefono':
                            df_filtrado['Telefono'] = df_filtrado['Telefono'].apply(normalizar_telefono_estricto)
                        else:
                            df_filtrado[col] = df_filtrado[col].astype(str).str.strip()
                    else:
                        df_filtrado[col] = ''
                
                df_filtrado['Docente'] = (df_filtrado['Docente']
                                          .str.replace(r'[\"\'\`]', '', regex=True)
                                          .str.replace(r'\s+', ' ', regex=True)
                                          .str.upper())
                
                if 'Programa_Origen' in df_filtrado.columns:
                    df_filtrado = df_filtrado.drop(columns=['Programa_Origen'])
                if 'Sede_Origen' in df_filtrado.columns:
                    df_filtrado = df_filtrado.drop(columns=['Sede_Origen'])
                    
                datos_consolidados.append(df_filtrado)
                time.sleep(0.01) 

    # 2. CONSOLIDACIÓN GLOBAL E INSERCIÓN DE ESTRUCTURA RECTORADO A-N
    if datos_consolidados:
        print("\n💾 Estructurando columnas finales y aplicando auditoría visual...")
        df_final = pd.concat(datos_consolidados, ignore_index=True)
        
        columnas_procesadas = ['Cedula', 'Docente', 'Correo', 'Telefono', 'Programa', 'Sede', 'Horas', 'Periodo', 'Observaciones']
        df_final = df_final.reindex(columns=columnas_procesadas)
        df_final.insert(0, 'Nº', range(1, len(df_final) + 1))
        
        df_final['CARGO'] = ''
        df_final['FECHA DE INGRESO'] = ''
        df_final['CONDICIÓN'] = ''
        df_final['OBSERVACIÓN RECTORADO'] = ''
        
        lista_cedulas = [c for c in df_final['Cedula'].tolist() if c != '']
        conteo_cedulas = Counter(lista_cedulas)
        
        nombre_salida = "Carga_Academica_Consolidada_Unificada.xlsx"
        
        with pd.ExcelWriter(nombre_salida, engine='openpyxl') as writer:
            df_final.to_excel(writer, sheet_name='Consolidado', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Consolidado']
            
            relleno_vacio_critico = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") 
            relleno_alerta_contacto = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") 
            relleno_duplicado = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") 
            fuente_duplicado = Font(color="9C0006", bold=True) 
            
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                
                if col_letter in ['K', 'L', 'M', 'N']:
                    worksheet.column_dimensions[col_letter].width = 24
                else:
                    worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
            for row_idx in range(2, worksheet.max_row + 1):
                celda_cedula = worksheet.cell(row=row_idx, column=2)
                celda_docente = worksheet.cell(row=row_idx, column=3)
                celda_telefono = worksheet.cell(row=row_idx, column=5) 
                celda_horas = worksheet.cell(row=row_idx, column=8)
                
                celda_telefono.number_format = '@'
                
                valor_cedula = str(celda_cedula.value).strip() if celda_cedula.value is not None else ""
                
                if valor_cedula == "":
                    celda_cedula.fill = relleno_vacio_critico
                elif conteo_cedulas[valor_cedula] > 1:
                    celda_cedula.fill = relleno_duplicado
                    celda_cedula.font = fuente_duplicado
                
                if celda_docente.value is None or str(celda_docente.value).strip() == "":
                    celda_docente.fill = relleno_vacio_critico
                    
                try:
                    val_horas = float(celda_horas.value) if celda_horas.value is not None else 0
                    if val_horas == 0:
                        celda_horas.fill = relleno_vacio_critico
                except ValueError:
                    celda_horas.fill = relleno_vacio_critico
                    
                for col_idx in [4, 5]:
                    celda = worksheet.cell(row=row_idx, column=col_idx)
                    if celda.value is None or str(celda.value).strip() == "" or str(celda.value).strip().lower() == "nan":
                        celda.fill = relleno_alerta_contacto
                        
        print(f"🚀 ¡ALGORITMO GLOBAL MULTITAREA DESPLEGADO!")
        print(f"• Soporte completo para programas compartidos (Ej: INGENIERÍA / PNFI).")
        print(f"• Soporte completo para sedes compartidas por fila (Ej: CABIMAS / CIUDAD OJEDA).")
        print(f"📍 Total registros unificados: {len(df_final)}")
        files.download(nombre_salida)
